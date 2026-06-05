"""Estado de Resultados Integral mensual.

Estructura del Excel esperado:
- Hoja 'EST. DE RES. MES A MES '
- Col 1 = código + descripción de cuenta (PUC, ej. '41502001 INTERESES FINANCIACION POLIZAS')
- Col 3..25 = valores mensuales (ENERO..DICIEMBRE), intervalos de 2
- Col 27 = ACUMULADO (no se usa, lo calculamos)
- Filas con prefijo 'Total ...' son subtotales (is_total=1)
- El año se infiere del título "ESTADO DE RESULTADO INTEGRAL - MENSUALES DEL <YYYY>"
"""
import io
import logging
import re
from fastapi import APIRouter, Depends, UploadFile, File, Request
import openpyxl
from app.database import get_db, get_connection
from app.auth.middleware import require_auth, require_superadmin
from app.sync.upload_helpers import upload_session

router = APIRouter(prefix="/api/estados-financieros", tags=["estados-financieros"])

_log = logging.getLogger("fide.financieros")
MAX_UPLOAD_MB = 25

# Columnas de meses en el Excel (0-indexed): ENERO en col 3, FEB en 5, ... DIC en 25
MONTH_COLS = {1: 3, 2: 5, 3: 7, 4: 9, 5: 11, 6: 13, 7: 15, 8: 17, 9: 19, 10: 21, 11: 23, 12: 25}


def _to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).strip()
    if not s: return 0.0
    try: return float(s.replace(',', '').replace('%', ''))
    except ValueError: return 0.0


def _parse_cuenta(raw: str):
    """De '41502001 INTERESES FINANCIACION POLIZAS' devuelve (code, desc, nivel, is_total).

    Detecta filas 'Total xxxx' como is_total=True con code=xxxx.
    También reconoce dos filas calculadas por el contador que NO tienen
    código numérico en el Excel (filas 221 y 233 de la plantilla):
      - 'UTILIDAD ANTES DE IMPUESTO DE RENTA' → code '539999' (entre 53 y 54)
      - 'UTILIDAD DEL EJERCICIO'              → code '999999' (al final)
    Estas filas llevan los valores REALES del contador (ya cuadrados, con
    decimales y signos correctos), no algo que reconstruyamos en backend.
    """
    raw = (raw or '').strip()
    if not raw:
        return None
    upper = raw.upper()
    if upper == 'CUENTA':
        return None
    # Líneas conclusión del Estado de Resultados sin código numérico.
    # Se les asigna un code sintético para que entren al pipeline normal.
    if 'UTILIDAD' in upper and 'ANTES DE IMPUESTO' in upper:
        return '539999', 'UTILIDAD ANTES DE IMPUESTO DE RENTA', 1, None, 1
    if 'UTILIDAD' in upper and ('DEL EJERCICIO' in upper or 'PERDIDA' in upper and 'EJERCICIO' in upper):
        return '999999', 'UTILIDAD (PERDIDA) DEL EJERCICIO', 1, None, 1
    # Detección de 'Total ...'
    is_total = raw.lower().startswith('total')
    if is_total:
        # 'Total 4150 INGRESOS FINANCIEROS' → code 4150
        m = re.match(r'Total\s+(\d+)\s*(.*)', raw, re.IGNORECASE)
        if not m:
            return None
        code = m.group(1)
        desc = m.group(2).strip()
    else:
        # Cuenta normal: '41502001 INTERESES ...'
        m = re.match(r'(\d+)\s+(.+)', raw)
        if not m:
            return None
        code = m.group(1)
        desc = m.group(2).strip()
    nivel = len(code)
    # Parent code: nivel anterior (cortar de a 2 dígitos para subgrupos PUC)
    parent = None
    if nivel >= 8:    parent = code[:6]
    elif nivel >= 6:  parent = code[:4]
    elif nivel >= 4:  parent = code[:2]
    elif nivel >= 2:  parent = code[:1]
    return code, desc, nivel, parent, is_total


def _detect_year(rows: list[tuple]) -> int | None:
    """Busca el año en las primeras filas del Excel."""
    for r in rows[:10]:
        for v in r:
            if v and isinstance(v, str):
                m = re.search(r'(20\d{2})', v)
                if m:
                    return int(m.group(1))
    return None


@router.post("/upload")
async def upload(request: Request, user=Depends(require_superadmin), file: UploadFile = File(...)):
    async with upload_session(
        request, user, file, source="financieros_upload", max_mb=MAX_UPLOAD_MB,
        generic_error_msg="No se pudo importar el archivo. Revisa el formato del Estado de Resultados."
    ) as ctx:
        # El Excel de Estados de Resultados usa una hoja específica (no la activa)
        wb = openpyxl.load_workbook(io.BytesIO(ctx.content), data_only=True, read_only=True)
        sheet_name = next((s for s in wb.sheetnames if 'mes' in s.lower() or 'res' in s.lower()), wb.sheetnames[0])
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        year = _detect_year(rows)
        if not year:
            raise ValueError("No se pudo detectar el año en el archivo")

        # Acumular por (cuenta_code, month) — el Excel tiene filas con códigos
        # repetidos (ej. 'Total 41 OPERACIONALES' aparece varias veces).
        # Sumamos los valores en lugar de fallar por UNIQUE constraint.
        agg = {}  # (code, month) -> {valor, desc, nivel, parent, is_total}
        # Detectamos meses PRESENTES en el archivo independiente del valor (puede
        # ser 0). Esto evita el bug detectado por la auditoría: si una cuenta cae
        # a 0 entre re-uploads, el DELETE no la limpiaba porque construíamos
        # months_in_file desde 'records' (que excluye valores 0).
        months_present_in_file = set()
        for r in rows:
            if not r or len(r) < 2:
                continue
            parsed = _parse_cuenta(r[1] if r[1] else '')
            if not parsed:
                continue
            code, desc, nivel, parent, is_total = parsed
            for month, col_idx in MONTH_COLS.items():
                if col_idx >= len(r):
                    continue
                raw_cell = r[col_idx]
                # Si la celda NO está vacía (incluye 0), el mes está presente en el archivo
                if raw_cell is not None and str(raw_cell).strip() != '':
                    months_present_in_file.add(month)
                val = _to_float(raw_cell)
                if val == 0:
                    continue
                key = (code, month)
                if key not in agg:
                    agg[key] = {
                        'descripcion': desc, 'nivel': nivel, 'parent': parent,
                        'is_total': 1 if is_total else 0, 'valor': 0.0,
                    }
                # Subtotales duplicados (mismo 'Total xxxx' repetido) → tomar el último valor
                # Cuentas detalle (is_total=0) → sumar (puede haber legit varias filas
                # de la misma cuenta auxiliar en distintos renglones).
                if is_total:
                    agg[key]['valor'] = val
                else:
                    agg[key]['valor'] += val

        # ROLLUP de categorías principales (códigos de 1 dígito) que aparecen
        # como header en el Excel pero NO tienen 'Total X' explícito.
        # Caso real: '5 GASTOS' en fila 61 es solo label; el contador no
        # totalizó al nivel 1 (solo Total 51, 52, 53, 54). Sin este rollup,
        # la cascada no muestra '5 GASTOS' como categoría principal del E.R.
        nivel1_headers = {}
        for r in rows:
            if not r or len(r) < 2:
                continue
            raw = str(r[1]).strip() if r[1] else ''
            if not raw or raw.lower().startswith('total'):
                continue
            m1 = re.match(r'^(\d)\s+(.+)', raw)
            if m1 and m1.group(1) not in nivel1_headers:
                nivel1_headers[m1.group(1)] = m1.group(2).strip()
        for code1, desc1 in nivel1_headers.items():
            # Si el contador YA totalizó este nivel (ej. 'Total 4 INGRESOS' →
            # code '4' en agg), no sobrescribir.
            if any(k[0] == code1 for k in agg.keys()):
                continue
            # Sumar descendientes detalle (is_total=0) cuyo código empiece con code1.
            for month in months_present_in_file:
                total = sum(
                    v['valor'] for (c, mm), v in agg.items()
                    if mm == month and not v['is_total'] and c.startswith(code1)
                )
                if total != 0:
                    agg[(code1, month)] = {
                        'descripcion': desc1, 'nivel': 1, 'parent': None,
                        'is_total': 1, 'valor': total,
                    }

        records = [
            {
                'year': year, 'month': m, 'cuenta_code': code,
                'cuenta_descripcion': v['descripcion'], 'nivel': v['nivel'],
                'parent_code': v['parent'], 'is_total': v['is_total'], 'valor': v['valor'],
            }
            for (code, m), v in agg.items()
        ]

        # Reemplazar meses presentes en el archivo (incluyendo aquellos con todas
        # las celdas en 0 — el DELETE limpia el mes para que las cuentas que
        # caían a 0 no queden con su valor anterior).
        months_in_file = sorted(months_present_in_file)

        with get_db() as conn:
            # Reemplazar SOLO los meses presentes en el archivo. Si el usuario
            # sube un archivo solo con la columna ABRIL → solo se reemplaza
            # abril; los meses previos permanecen intactos. Si sube uno
            # acumulado (Ene-Abr) → se reemplazan los 4.
            if months_in_file:
                placeholders = ','.join(['?'] * len(months_in_file))
                conn.execute(
                    f"DELETE FROM estados_financieros WHERE year = ? AND month IN ({placeholders})",
                    [year, *months_in_file]
                )
            for rec in records:
                conn.execute("""
                    INSERT INTO estados_financieros
                    (year, month, cuenta_code, cuenta_descripcion, nivel, parent_code, is_total, valor, sync_batch_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (rec['year'], rec['month'], rec['cuenta_code'], rec['cuenta_descripcion'],
                      rec['nivel'], rec['parent_code'], rec['is_total'], rec['valor'], ctx.sync_id))

        nombres_mes = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        meses_str = ', '.join(nombres_mes[m] for m in months_in_file)
        ctx.set_counts(fetched=len(rows), inserted=len(records))
        ctx.set_audit_extra(f"year={year} meses=[{meses_str}]")
        return {
            "status": "success",
            "year": year,
            "meses_cargados": meses_str,
            "n_meses": len(months_in_file),
            "records": len(records),
        }


@router.get("/years")
def years_available(_user=Depends(require_auth)):
    conn = get_connection()
    try:
        rows = conn.execute("SELECT DISTINCT year FROM estados_financieros ORDER BY year DESC").fetchall()
        return [r["year"] for r in rows]
    finally:
        conn.close()


@router.get("")
def list_records(year: int, _user=Depends(require_auth)):
    """Todas las filas del año (todos los meses y cuentas), incluyendo subtotales."""
    conn = get_connection()
    try:
        rows = conn.execute(
            "SELECT year, month, cuenta_code, cuenta_descripcion, nivel, parent_code, is_total, valor "
            "FROM estados_financieros WHERE year = ? ORDER BY cuenta_code, month",
            (year,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@router.get("/summary")
def summary(year: int, _user=Depends(require_auth)):
    """KPIs del año: ingresos, gastos, costos, utilidad bruta/operacional/neta + por mes."""
    conn = get_connection()
    try:
        # Agregados por mes a partir de la base. NO recalculamos la utilidad —
        # usamos las filas REALES del Excel (códigos sintéticos 539999 y 999999),
        # que el contador ya cuadró con sus decimales y signos.
        # gastos_op excluye 54 (impuesto de renta) que aparece después en el ER.
        agg = conn.execute("""
            SELECT
                month,
                SUM(CASE WHEN substr(cuenta_code,1,1)='4' AND is_total=0 THEN valor ELSE 0 END) as ingresos,
                SUM(CASE WHEN substr(cuenta_code,1,1)='5'
                          AND substr(cuenta_code,1,2)!='54'
                          AND cuenta_code NOT IN ('539999','999999')
                          AND is_total=0 THEN valor ELSE 0 END) as gastos_op,
                SUM(CASE WHEN substr(cuenta_code,1,2)='54' AND is_total=0 THEN valor ELSE 0 END) as impuesto_renta,
                SUM(CASE WHEN substr(cuenta_code,1,1)='6' AND is_total=0 THEN valor ELSE 0 END) as costos,
                SUM(CASE WHEN substr(cuenta_code,1,1)='7' AND is_total=0 THEN valor ELSE 0 END) as costos_prod,
                SUM(CASE WHEN cuenta_code='539999' THEN valor ELSE 0 END) as utilidad_antes_excel,
                SUM(CASE WHEN cuenta_code='999999' THEN valor ELSE 0 END) as utilidad_neta_excel
            FROM estados_financieros WHERE year = ?
            GROUP BY month ORDER BY month
        """, (year,)).fetchall()

        by_month = []
        total_ingresos = total_gastos_op = total_impuesto = total_costos = 0.0
        total_util_antes_excel = total_util_neta_excel = 0.0
        for r in agg:
            d = dict(r)
            d['gastos'] = d['gastos_op'] or 0
            # Si el Excel trae la fila calculada del contador, usar ESA. Si no
            # (archivo viejo, o el contador no la diligenció), recalcular.
            ua_excel = d.get('utilidad_antes_excel') or 0
            un_excel = d.get('utilidad_neta_excel') or 0
            if ua_excel != 0:
                d['utilidad_antes_impuesto'] = ua_excel
            else:
                d['utilidad_antes_impuesto'] = (d['ingresos'] or 0) - d['gastos'] - (d['costos'] or 0) - (d['costos_prod'] or 0)
            if un_excel != 0:
                d['utilidad'] = un_excel
            else:
                d['utilidad'] = d['utilidad_antes_impuesto'] - (d['impuesto_renta'] or 0)
            by_month.append(d)
            total_ingresos += d['ingresos'] or 0
            total_gastos_op += d['gastos']
            total_impuesto += d['impuesto_renta'] or 0
            total_costos += (d['costos'] or 0) + (d['costos_prod'] or 0)
            total_util_antes_excel += ua_excel
            total_util_neta_excel += un_excel
        # Totales: si el Excel trae las filas, sumamos las del contador (cada mes).
        # Si no, calculamos. Esto soporta archivos parciales o legacy.
        if total_util_antes_excel != 0:
            utilidad_antes = total_util_antes_excel
        else:
            utilidad_antes = total_ingresos - total_gastos_op - total_costos
        if total_util_neta_excel != 0:
            utilidad_neta = total_util_neta_excel
        else:
            utilidad_neta = utilidad_antes - total_impuesto
        margen = (utilidad_neta / total_ingresos * 100) if total_ingresos > 0 else 0
        margen_antes = (utilidad_antes / total_ingresos * 100) if total_ingresos > 0 else 0
        n_months = len(by_month)
        return {
            'year': year,
            'meses_con_datos': n_months,
            'ingresos_total': total_ingresos,
            'gastos_total': total_gastos_op,
            'impuesto_renta_total': total_impuesto,
            'costos_total': total_costos,
            'utilidad_antes_impuesto_total': utilidad_antes,
            'utilidad_total': utilidad_neta,
            'margen_antes_impuesto_pct': margen_antes,
            'margen_pct': margen,
            'ingresos_promedio_mensual': total_ingresos / n_months if n_months else 0,
            'gastos_promedio_mensual': total_gastos_op / n_months if n_months else 0,
            'utilidad_promedio_mensual': utilidad_neta / n_months if n_months else 0,
            'by_month': by_month,
        }
    finally:
        conn.close()
