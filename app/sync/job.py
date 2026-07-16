"""Sync job: imports data from an uploaded Excel file (processed in-memory)."""
import time
from datetime import datetime
from app.database import get_db, CREDIT_FIELDS
from app.sync.upload_helpers import check_file_signature, _parse_with_zip_xml

MAX_ROWS = 50_000

# Firma del Informe de Cartera (los 3 exports — vieja, nueva, consolidado —
# comparten estos headers; verificado contra los .xlsx reales). Sin esta
# validación, subir el archivo equivocado por el botón de cartera respondía
# "success" mientras reemplazaba la cartera completa con basura (cédulas como
# estados, $132 billones de desembolso) — el peor fallo posible: silencioso.
CARTERA_REQUIRE = ('cuenta', 'identificacion', 'cliente', 'estado',
                   'saldo_capital', 'valor_cuota')
CARTERA_REJECT = (
    ('tipo mvto', 'Recaudo/Movimientos'),
    ('fecha movimiento', 'Recaudo/Movimientos'),
    ('paso ruta', 'Solicitudes'),
    ('saldo vigente', 'Resumen Estado Cuenta (Saldo Cartera)'),
    ('naturaleza del litigio', 'Procesos Judiciales'),
)


def _read_cartera_rows(file_bytes: bytes) -> tuple[list, str]:
    """Lee el Excel de cartera devolviendo (todas las filas, nombre de hoja).

    Mantiene la selección de hoja 'Cartera_Consolidado' cuando existe. Si
    openpyxl revienta por estilos corruptos (defecto conocido de algunos
    exports de la plataforma), cae al parser XML directo — el mismo fallback
    que ya usan los demás módulos vía read_excel().
    """
    import openpyxl
    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet_name = "Cartera_Consolidado" if "Cartera_Consolidado" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows, sheet_name
    except Exception:
        return _parse_with_zip_xml(file_bytes), "(hoja 1, parser XML)"


def _validated_cartera_rows(file_bytes: bytes) -> tuple[list, list, str]:
    """Lee + valida firma. Devuelve (fila de encabezados, filas de datos, hoja).

    El encabezado se devuelve aparte porque el transformer resuelve los índices
    de columna POR NOMBRE a partir de él (inmune a que la plataforma inserte
    columnas)."""
    all_rows, sheet_name = _read_cartera_rows(file_bytes)
    hdr_idx = check_file_signature(
        all_rows, archivo="Informe de Cartera",
        require=CARTERA_REQUIRE, reject=CARTERA_REJECT,
    )
    header = all_rows[hdr_idx]
    rows = all_rows[hdr_idx + 1:]
    if len(rows) > MAX_ROWS:
        raise ValueError(f"Archivo excede el límite de {MAX_ROWS} filas")
    if not rows:
        raise ValueError(
            "El archivo tiene los encabezados correctos pero no trae filas de "
            "datos. No se importó nada — la cartera actual queda intacta."
        )
    return header, rows, sheet_name

# Campos que se actualizan desde el archivo de la plataforma NUEVA (cartera nueva)
# sobre la cartera BASE (cartera vieja). El resto de campos se preserva del base.
# Corresponde a las columnas naranja del 'Informe cartera plataforma vieja.xlsx'.
ORANGE_FIELDS = (
    'estado', 'saldo_capital', 'saldo_favor', 'valor_cuota', 'fecha_ult_pago',
    'calificacion', 'cuotas_pagadas', 'dias_mora', 'maxima_mora', 'aliado',
)


def _insert_credits(conn, records: list[dict], sync_id: int):
    """Inserta créditos con identificacion/cliente cifrados + sus masks
    pre-calculadas (acelera /api/credits al no requerir decrypt cada lectura)."""
    from app.crypto import decrypt, mask_identificacion, mask_cliente
    extra_cols = ["sync_batch_id", "identificacion_masked", "cliente_masked"]
    placeholders = ", ".join(["?"] * (len(CREDIT_FIELDS) + len(extra_cols)))
    cols = ", ".join(CREDIT_FIELDS + extra_cols)
    for record in records:
        # Las masks se computan desde el cifrado (no tenemos plaintext aquí)
        im = mask_identificacion(decrypt(record.get("identificacion")))
        cm = mask_cliente(decrypt(record.get("cliente")))
        values = [record.get(k) for k in CREDIT_FIELDS] + [sync_id, im, cm]
        conn.execute(f"INSERT INTO credits ({cols}) VALUES ({placeholders})", values)


def _cleanup_old_batches(conn, current_sync_id: int):
    """Mantiene los últimos 7 días de datos de cartera (manual_upload) y 30
    días de sync_logs. Filtra estrictamente por source='manual_upload' para
    evitar eliminar datos de otros módulos (recaudo, solicitudes, etc.)."""
    conn.execute("""
        DELETE FROM credits WHERE sync_batch_id IN (
            SELECT id FROM sync_logs
            WHERE id != ?
              AND source = 'manual_upload'
              AND started_at < datetime('now', '-7 days')
        )
    """, (current_sync_id,))
    conn.execute(
        "DELETE FROM sync_logs WHERE source = 'manual_upload' "
        "AND started_at < datetime('now', '-30 days') AND status != 'running'"
    )


def sync_from_excel(file_bytes: bytes, uploaded_by: int | None = None) -> dict:
    """Import cartera data from an in-memory Excel file. Raises on failure."""
    from app.acano.transformer import transform_excel_batch

    # Create the sync_log first so failures are recorded.
    with get_db() as conn:
        conn.execute(
            "INSERT INTO sync_logs (started_at, status, source, uploaded_by) "
            "VALUES (?, 'running', 'manual_upload', ?)",
            (datetime.utcnow().isoformat(), uploaded_by)
        )
        sync_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    start = time.time()
    try:
        header, rows, sheet_name = _validated_cartera_rows(file_bytes)

        records = transform_excel_batch(rows, header)

        # Skip rows with no identificacion/cliente/estado — usually empty trailing rows.
        records = [r for r in records
                   if r.get("identificacion") and r.get("cliente") and r.get("estado")]
        if not records:
            # Sin este guard, un archivo "válido pero vacío" reemplazaba el
            # batch activo con nada y el dashboard quedaba en blanco con un
            # mensaje de éxito.
            raise ValueError(
                "El archivo no trae créditos válidos (0 filas con "
                "identificación, cliente y estado). No se importó nada — la "
                "cartera actual queda intacta."
            )

        with get_db() as conn:
            _insert_credits(conn, records, sync_id)
            _cleanup_old_batches(conn, sync_id)
            # Reload base = perdemos tracking de plataforma nueva.
            # El próximo incremental_update_from_excel lo repuebla.
            conn.execute("DELETE FROM cartera_nueva_cuentas")
            duration = time.time() - start
            conn.execute("""
                UPDATE sync_logs SET status='success', completed_at=?,
                records_fetched=?, records_inserted=?, duration_seconds=?
                WHERE id=?
            """, (datetime.utcnow().isoformat(), len(rows), len(records), duration, sync_id))

        return {
            "status": "success",
            "records": len(records),
            "sheet": sheet_name,
            "duration": round(duration, 2)
        }

    except Exception as e:
        duration = time.time() - start
        try:
            with get_db() as conn:
                conn.execute("""
                    UPDATE sync_logs SET status='failed', completed_at=?,
                    error_message=?, duration_seconds=? WHERE id=?
                """, (datetime.utcnow().isoformat(), f"{type(e).__name__}: {str(e)[:400]}",
                      duration, sync_id))
        except Exception:
            pass
        raise


def incremental_update_from_excel(file_bytes: bytes, uploaded_by: int | None = None) -> dict:
    """Actualización incremental de cartera desde el archivo de la plataforma NUEVA.

    Política:
      - Para cada cuenta presente en el archivo nuevo:
          * Si ya existe en la base → UPDATE solo las columnas naranja (ORANGE_FIELDS).
          * Si no existe → INSERT como crédito nuevo.
      - Cuentas existentes que NO están en el archivo nuevo permanecen intactas.

    Se crea un nuevo sync_log y todos los créditos del batch anterior se
    migran a este nuevo batch (para que las queries del dashboard, que
    apuntan al último 'manual_upload' exitoso, sigan viendo todo).
    """
    from app.acano.transformer import transform_excel_batch

    with get_db() as conn:
        conn.execute(
            "INSERT INTO sync_logs (started_at, status, source, uploaded_by) "
            "VALUES (?, 'running', 'manual_upload', ?)",
            (datetime.utcnow().isoformat(), uploaded_by)
        )
        sync_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    start = time.time()
    try:
        header, rows, sheet_name = _validated_cartera_rows(file_bytes)

        new_records = transform_excel_batch(rows, header)
        # Necesitamos cuenta como clave de matching; descartamos filas vacías.
        new_records = [r for r in new_records if r.get("cuenta") and r.get("identificacion")]
        if not new_records:
            raise ValueError(
                "El archivo no trae créditos válidos (0 filas con cuenta e "
                "identificación). No se actualizó nada — la cartera queda intacta."
            )

        updated_count = 0
        inserted_count = 0

        with get_db() as conn:
            # Migrar todos los créditos del último batch exitoso anterior al nuevo batch,
            # para preservar el base + actualizaciones previas en el batch activo.
            prev_sync = conn.execute(
                "SELECT id FROM sync_logs WHERE source='manual_upload' AND status='success' "
                "AND id < ? ORDER BY id DESC LIMIT 1", (sync_id,)
            ).fetchone()
            if prev_sync:
                conn.execute(
                    "UPDATE credits SET sync_batch_id=? WHERE sync_batch_id=?",
                    (sync_id, prev_sync[0])
                )

            # INSERT también popula las masks pre-computadas (acelera /api/credits).
            from app.crypto import decrypt as _dec, mask_identificacion as _mi, mask_cliente as _mc
            extra_cols = ["sync_batch_id", "identificacion_masked", "cliente_masked"]
            placeholders = ", ".join(["?"] * (len(CREDIT_FIELDS) + len(extra_cols)))
            cols_sql = ", ".join(CREDIT_FIELDS + extra_cols)

            touched_cuentas = []
            for rec in new_records:
                cuenta = rec.get('cuenta')
                if not cuenta:
                    continue
                touched_cuentas.append(str(cuenta))
                existing = conn.execute(
                    "SELECT id FROM credits WHERE cuenta=? AND sync_batch_id=?",
                    (str(cuenta), sync_id)
                ).fetchone()
                if existing:
                    # UPDATE solo campos naranja con valor no-None del nuevo archivo
                    update_pairs = [(f, rec[f]) for f in ORANGE_FIELDS
                                    if f in rec and rec[f] is not None]
                    if update_pairs:
                        set_clause = ', '.join(f"{f}=?" for f, _ in update_pairs)
                        params = [v for _, v in update_pairs] + [existing[0]]
                        conn.execute(f"UPDATE credits SET {set_clause} WHERE id=?", params)
                        updated_count += 1
                else:
                    im = _mi(_dec(rec.get("identificacion")))
                    cm = _mc(_dec(rec.get("cliente")))
                    values = [rec.get(k) for k in CREDIT_FIELDS] + [sync_id, im, cm]
                    conn.execute(f"INSERT INTO credits ({cols_sql}) VALUES ({placeholders})", values)
                    inserted_count += 1

            # Repoblar tabla de tracking de plataforma. Las cuentas que aparecen
            # en este archivo se marcan como "nueva"; el resto que ya está en
            # credits queda implícitamente como "vieja".
            conn.execute("DELETE FROM cartera_nueva_cuentas")
            for c in touched_cuentas:
                conn.execute(
                    "INSERT OR REPLACE INTO cartera_nueva_cuentas (cuenta, last_seen_at) "
                    "VALUES (?, datetime('now'))",
                    (c,)
                )

            _cleanup_old_batches(conn, sync_id)

            duration = time.time() - start
            conn.execute(
                "UPDATE sync_logs SET status='success', completed_at=?, "
                "records_fetched=?, records_inserted=?, duration_seconds=? WHERE id=?",
                (datetime.utcnow().isoformat(), len(rows),
                 updated_count + inserted_count, duration, sync_id)
            )

        return {
            "status": "success",
            "actualizados": updated_count,
            "nuevos": inserted_count,
            "filas_archivo": len(rows),
            "duration": round(duration, 2),
        }

    except Exception as e:
        duration = time.time() - start
        try:
            with get_db() as conn:
                conn.execute(
                    "UPDATE sync_logs SET status='failed', completed_at=?, "
                    "error_message=?, duration_seconds=? WHERE id=?",
                    (datetime.utcnow().isoformat(),
                     f"{type(e).__name__}: {str(e)[:400]}", duration, sync_id)
                )
        except Exception:
            pass
        raise
